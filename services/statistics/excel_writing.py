import openpyxl


def create_excel_table(data):
    # Создаем новую книгу (Excel файл)
    wb = openpyxl.Workbook()

    # Выбираем активный лист
    sheet = wb.active

    # Записываем данные в ячейки
    sheet['A1'] = 'Канал'
    sheet['B1'] = 'Просмотры'
    sheet['C1'] = 'Репосты'
    row = 2

    for item in data:
        channel_id = str(item['id'])
        views = item['stat']['views']
        forwards = item['stat']['forwards']

        sheet.cell(row=row, column=1, value=channel_id)
        sheet.cell(row=row, column=2, value=views)
        sheet.cell(row=row, column=3, value=forwards)

        row += 1

    # Сохраняем книгу под заданным именем
    wb.save('table.xlsx')
